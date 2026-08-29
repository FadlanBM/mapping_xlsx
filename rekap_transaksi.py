import os
import re
import random
from typing import Optional, List, Tuple
import pandas as pd
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows

TARGET_COLUMNS = ['TANGGAL', 'NAMA AKUN', 'BUKTI', 'DEBET', 'KREDIT', 'KETERANGAN']


def _clean_numeric(val) -> float:
    """Membersihkan dan mengkonversi format angka/nominal menjadi float."""
    if pd.isna(val) or val is None:
        return 0.0
    if isinstance(val, (int, float)):
        return float(val)
    val_str = str(val).strip().replace("Rp", "").replace("rp", "").replace("RP", "").strip()
    if not val_str or val_str == "-" or val_str == "nan" or val_str == "None":
        return 0.0
    
    # Handle separators
    if "." in val_str and "," in val_str:
        # Format Indonesia: 1.000.000,50 -> 1000000.50
        val_str = val_str.replace(".", "").replace(",", ".")
    elif "." in val_str:
        parts = val_str.split(".")
        if len(parts[-1]) == 3 and len(parts) > 1:
            val_str = val_str.replace(".", "")
    elif "," in val_str:
        parts = val_str.split(",")
        if len(parts[-1]) == 3 and len(parts) > 1:
            val_str = val_str.replace(",", "")
        else:
            val_str = val_str.replace(",", ".")
            
    try:
        return float(val_str)
    except ValueError:
        return 0.0


def detect_and_load_data(file_path: str) -> pd.DataFrame:
    """
    Membaca file Excel / CSV dengan mendeteksi baris header secara otomatis
    meskipun terdapat baris kosong atau judul di baris-baris awal.
    """
    if not os.path.exists(file_path):
        raise FileNotFoundError(f"File '{file_path}' tidak ditemukan.")

    is_excel = file_path.lower().endswith(('.xlsx', '.xls', '.xlsm'))
    
    if is_excel:
        # Baca 20 baris pertama untuk mencari baris header
        preview_df = pd.read_excel(file_path, header=None, nrows=20)
    else:
        preview_df = pd.read_csv(file_path, header=None, nrows=20)

    header_idx = 0
    max_matches = 0
    
    # Kata kunci penanda kolom
    keywords = ['BUKTI', 'AKUN', 'DEBET', 'DEBIT', 'KREDIT', 'CREDIT', 'KETERANGAN', 'TANGGAL']
    
    for idx, row in preview_df.iterrows():
        row_str = " ".join([str(val).upper() for val in row if pd.notna(val)])
        matches = sum(1 for kw in keywords if kw in row_str)
        if matches > max_matches:
            max_matches = matches
            header_idx = idx

    # Baca file dengan header_idx yang ditemukan
    if is_excel:
        df = pd.read_excel(file_path, header=header_idx)
    else:
        df = pd.read_csv(file_path, header=header_idx)

    # Filter baris yang benar-benar kosong
    df = df.dropna(how='all')
    return df


def clean_and_standardize(df: pd.DataFrame) -> pd.DataFrame:
    """
    Menyelaraskan nama kolom ke: TANGGAL (opsional), NAMA AKUN, BUKTI, DEBET, KREDIT, KETERANGAN.
    Mengkonversi dan merapikan tipe data.
    """
    df = df.copy()
    col_mapping = {}
    
    for col in df.columns:
        clean_col = re.sub(r'[^a-zA-Z0-9]', '', str(col).strip().upper())
        if 'TANGGAL' in clean_col or 'DATE' in clean_col or 'TGL' in clean_col:
            col_mapping[col] = 'TANGGAL'
        elif 'AKUN' in clean_col or 'ACCOUNT' in clean_col or 'REKENING' in clean_col or 'PERKIRAAN' in clean_col:
            col_mapping[col] = 'NAMA AKUN'
        elif 'BUKTI' in clean_col or 'NOBUKTI' in clean_col or 'VOUCHER' in clean_col or 'INVOICE' in clean_col or 'TRX' in clean_col or 'ID' in clean_col:
            col_mapping[col] = 'BUKTI'
        elif 'DEBET' in clean_col or 'DEBIT' in clean_col:
            col_mapping[col] = 'DEBET'
        elif 'KREDIT' in clean_col or 'CREDIT' in clean_col:
            col_mapping[col] = 'KREDIT'
        elif any(k in clean_col for k in ['KETERANGAN', 'KET', 'DESC', 'DESKRIPSI', 'CATATAN', 'MEMO']):
            col_mapping[col] = 'KETERANGAN'

    df.rename(columns=col_mapping, inplace=True)
    
    # Kolom inti yang harus ada
    required_cols = ['NAMA AKUN', 'BUKTI', 'DEBET', 'KREDIT', 'KETERANGAN']
    for col in required_cols:
        if col not in df.columns:
            df[col] = 0.0 if col in ['DEBET', 'KREDIT'] else ""

    # Pembersihan tipe data
    df['DEBET'] = df['DEBET'].apply(_clean_numeric)
    df['KREDIT'] = df['KREDIT'].apply(_clean_numeric)
    df['BUKTI'] = df['BUKTI'].fillna('').astype(str).str.strip()
    df['NAMA AKUN'] = df['NAMA AKUN'].fillna('').astype(str).str.strip()
    df['KETERANGAN'] = df['KETERANGAN'].fillna('').astype(str).str.strip()

    # Bersihkan baris 'nan' string atau baris kosong tak berguna
    df = df[~((df['BUKTI'].isin(['', 'nan', 'None', 'NaN'])) & (df['DEBET'] == 0) & (df['KREDIT'] == 0))]

    # Tentukan urutan kolom akhir
    if 'TANGGAL' in df.columns:
        def format_date_val(v):
            if pd.isna(v) or str(v).strip() in ['', 'nan', 'NaT']:
                return ""
            if isinstance(v, pd.Timestamp):
                return v.strftime('%Y-%m-%d')
            try:
                dt = pd.to_datetime(v)
                return dt.strftime('%Y-%m-%d')
            except Exception:
                return str(v).strip()
                
        df['TANGGAL'] = df['TANGGAL'].apply(format_date_val)
        final_cols = ['TANGGAL', 'NAMA AKUN', 'BUKTI', 'DEBET', 'KREDIT', 'KETERANGAN']
    else:
        final_cols = ['NAMA AKUN', 'BUKTI', 'DEBET', 'KREDIT', 'KETERANGAN']

    return df[final_cols]


def sort_by_bukti_and_type(df: pd.DataFrame, debet_first: bool = True) -> pd.DataFrame:
    """
    Mengelompokkan data berdasarkan nomor BUKTI dan mengurutkan baris transaksi
    pada BUKTI yang sama sehingga baris DEBET dan KREDIT tersusun secara berurutan.
    """
    df = clean_and_standardize(df)
    
    # Menentukan prioritas baris dalam BUKTI yang sama:
    # Debet (DEBET > 0) -> 1, Kredit (KREDIT > 0) -> 2 jika debet_first=True
    def calculate_priority(row):
        d_val = row['DEBET']
        k_val = row['KREDIT']
        if d_val > 0 and k_val == 0:
            return 1 if debet_first else 2
        elif k_val > 0 and d_val == 0:
            return 2 if debet_first else 1
        elif d_val > 0 and k_val > 0:
            return 1.5
        return 3

    df['_sort_priority'] = df.apply(calculate_priority, axis=1)
    
    # Urutkan berdasarkan nomor BUKTI, kemudian prioritas debet/kredit
    sort_keys = ['BUKTI', '_sort_priority']
    df_sorted = df.sort_values(
        by=sort_keys, 
        ascending=[True, True],
        kind='mergesort' # preserve original row order for equal priority
    ).copy()
    
    df_sorted.drop(columns=['_sort_priority'], inplace=True)
    return df_sorted.reset_index(drop=True)


def export_to_excel(df: pd.DataFrame, output_path: str, add_totals: bool = False) -> str:
    """
    Mengekspor DataFrame ke file Excel dengan performa tinggi & styling profesional.
    Otomatis menambahkan ekstensi .xlsx jika belum ada.
    """
    if not output_path.lower().endswith('.xlsx'):
        base, _ = os.path.splitext(output_path)
        output_path = f"{base}.xlsx"

    wb = Workbook()
    ws = wb.active
    ws.title = "Rekap Transaksi"
    
    # Pastikan gridlines tetap terlihat
    ws.views.sheetView[0].showGridLines = True

    # Styling Palettes
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid") # Dark Navy
    header_font = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
    
    row_font = Font(name="Segoe UI", size=10)
    total_font = Font(name="Segoe UI", size=10, bold=True)
    
    thin_side = Side(style="thin", color="E0E0E0")
    thin_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
    
    double_bottom_side = Side(style="double", color="000000")
    total_top_side = Side(style="thin", color="000000")
    total_border = Border(top=total_top_side, bottom=double_bottom_side, left=thin_side, right=thin_side)

    # Siapkan data frame final (tambahkan baris total jika diminta)
    export_df = df.copy()
    has_total_row = False
    if add_totals and len(export_df) > 0:
        total_debet = export_df['DEBET'].sum()
        total_kredit = export_df['KREDIT'].sum()
        total_dict = {col: '' for col in export_df.columns}
        total_dict['NAMA AKUN'] = 'TOTAL KESELURUHAN'
        total_dict['DEBET'] = total_debet
        total_dict['KREDIT'] = total_kredit
        export_df = pd.concat([export_df, pd.DataFrame([total_dict])], ignore_index=True)
        has_total_row = True

    # Tulis semua baris sekaligus menggunakan dataframe_to_rows (sangat cepat)
    for r in dataframe_to_rows(export_df, index=False, header=True):
        ws.append(r)

    headers = list(export_df.columns)
    debet_col_idx = headers.index('DEBET') + 1 if 'DEBET' in headers else -1
    kredit_col_idx = headers.index('KREDIT') + 1 if 'KREDIT' in headers else -1
    bukti_col_idx = headers.index('BUKTI') + 1 if 'BUKTI' in headers else -1
    tgl_col_idx = headers.index('TANGGAL') + 1 if 'TANGGAL' in headers else -1

    # Format Header (Row 1)
    ws.row_dimensions[1].height = 25
    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border

    # Format Data Cells
    total_row_idx = ws.max_row if has_total_row else -1
    num_format = '#,##0.00;(#,##0.00);"-";@'

    align_center = Alignment(horizontal="center", vertical="center")
    align_right = Alignment(horizontal="right", vertical="center")
    align_left = Alignment(horizontal="left", vertical="center")

    for row_idx in range(2, ws.max_row + 1):
        is_total = (row_idx == total_row_idx)
        current_font = total_font if is_total else row_font
        current_border = total_border if is_total else thin_border

        for col_idx in range(1, len(headers) + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.font = current_font
            cell.border = current_border
            
            if col_idx in (debet_col_idx, kredit_col_idx):
                cell.number_format = num_format
                cell.alignment = align_right
            elif col_idx in (bukti_col_idx, tgl_col_idx):
                cell.alignment = align_center
            else:
                cell.alignment = align_left

    # Auto-fit column widths berdasarkan sample
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        sample_cells = col[:100]
        for cell in sample_cells:
            val_str = str(cell.value or '')
            if len(val_str) > max_len:
                max_len = len(val_str)
        ws.column_dimensions[col_letter].width = max(min(max_len + 4, 60), 14)

    os.makedirs(os.path.dirname(os.path.abspath(output_path)), exist_ok=True)
    
    try:
        wb.save(output_path)
    except PermissionError:
        # Jika file sedang dibuka oleh Excel, simpan dengan alternatif nama
        base, ext = os.path.splitext(output_path)
        alt_output = f"{base}_rekap{ext}"
        wb.save(alt_output)
        return alt_output
        
    return output_path


def generate_sample_excel(output_path: str, count: int = 5) -> str:
    """
    Membuat file Excel simulasi transaksi dengan kolom:
    'TANGGAL', 'NAMA AKUN', 'BUKTI', 'DEBET', 'KREDIT', 'KETERANGAN' dalam urutan acak.
    Otomatis menambahkan ekstensi .xlsx jika belum ada.
    """
    if not output_path.lower().endswith('.xlsx'):
        base, _ = os.path.splitext(output_path)
        output_path = f"{base}.xlsx"

    random.seed(123)
    sample_accounts = [
        ("Persediaan Barang", "Kas"),
        ("Beban Sewa Gedung", "Bank BCA"),
        ("Peralatan Kantor", "Utang Usaha"),
        ("Beban Listrik & Air", "Kas Kecil"),
        ("Pembelian Bahan Baku", "Bank Mandiri")
    ]
    
    rows = []
    for i in range(1, count + 1):
        bukti_no = f"BKT-{i:03d}"
        nominal = random.randint(25, 300) * 100000
        debet_acc, kredit_acc = sample_accounts[(i - 1) % len(sample_accounts)]
        tgl_str = f"2026-08-{i:02d}"
        
        debet_entry = {
            "TANGGAL": tgl_str,
            "NAMA AKUN": debet_acc,
            "BUKTI": bukti_no,
            "DEBET": nominal,
            "KREDIT": 0,
            "KETERANGAN": f"Transaksi {debet_acc} (#{i})"
        }
        kredit_entry = {
            "TANGGAL": tgl_str,
            "NAMA AKUN": kredit_acc,
            "BUKTI": bukti_no,
            "DEBET": 0,
            "KREDIT": nominal,
            "KETERANGAN": f"Pengeluaran via {kredit_acc} (#{i})"
        }
        
        pair = [debet_entry, kredit_entry]
        if random.choice([True, False]):
            pair.reverse()
        rows.extend(pair)
        
    random.shuffle(rows)
    df = pd.DataFrame(rows)
    os.makedirs(os.path.dirname(os.path.abspath(output_path)), exist_ok=True)
    df.to_excel(output_path, index=False)
    return output_path
